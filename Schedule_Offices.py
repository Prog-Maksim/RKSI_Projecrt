import googleapiclient
from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl
import schedule
import requests
from bs4 import BeautifulSoup
import telebot
from datetime import date, datetime
import re
from id_offices import name_teacher, CalendarID_group, Token_bot, TG_Bot_ID


# Создаем подключение к календарю указывая для него разрешение доступа
SCOPES = ['https://www.googleapis.com/auth/calendar']
calendarId = '35fc2b53f41517c857c7cd03942345f7ae6770f333eb04e36c768d03ce453e04@group.calendar.google.com'
SERVICE_ACCOUNT_FILE = 'TESTS.json'
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = googleapiclient.discovery.build('calendar', 'v3', credentials=credentials)

bot = telebot.TeleBot(Token_bot)
id_events = list()  # Храним id ивентов которые надо удалить
Name_list = list()  # Хранит список всех преподователей и групп из планшетки

def start_message(message: str, text: str):
    """Функция отправляет админам сообщения об ошибки"""
    for i in message:
        bot.send_message(i, f"{text}")

def Delete_Events():
    """Функция удаляет найденные ивенты в календаре"""
    num = -1  # Создаем переменную для поиска id в списке
    for i in id_events:  # Проходимся по списку
        try:
            num += 1  # Сразу же увеличиваем ее
            service.events().delete(calendarId=CalendarID_group[i], eventId=id_events[num + 1]).execute()  # Находим нужный
        except:
            pass

    id_events.clear()

def New_Events(kab, group, preps, time, day, name):
    if '-' not in time:
        time = time.split(' — ')
    else:
        time = time.split('-')

    time[0] = time[0].replace(' ', '')
    time[-1] = time[-1].replace(' ', '')
    event = {
        'summary': f'{group} — {preps}',
        'description': f'{group} — {preps}  ({name})',
        'colorId': '1',
        'start': {'dateTime': f'{day}T{time[0]}:00+03:00'},
        'end': {'dateTime': f'{day}T{time[-1]}:00+03:00'}
    }
    try:
        service.events().insert(calendarId=CalendarID_group[kab], body=event).execute()  # Выполняем запрос о создании ивента
    except Exception as e:
        start_message(TG_Bot_ID, f'''Произошла ошибка
                                 ошибка была вызвана этими данными:
                                 text = {group} — {preps}, \ndescription = {group} — {preps}, \nday = {day}, \ntime = {time}, \ncolor = 1, \nkab = {kab}
                                 описание ошибки: {e}''')

def clear_calendar():
    for i in CalendarID_group:
        now = datetime.utcnow().isoformat() + 'Z'
        events_result = service.events().list(calendarId=CalendarID_group[i], timeMin=now, maxResults=100, singleEvents=True, orderBy='startTime').execute()
        events = events_result.get('items', [])  # Ищем ивенты
        for event in events:  # Проходимся по найденным ивентам
            strs = event.get('id')  # Получаем id ивента
            print(f'{i} -- {strs}')
            id_events.append(i)  # Записываем в список у какой группы мы парсили календарь
            id_events.append(strs)  # добавляем в список id ивентов которые надо удалить
    Delete_Events()

def get_events_list(time: str()):  # Принимает аргумент дата из планшетки
    """Функция находит id ивентов в календаре и записывает их в лист"""
    id_events.clear()  # Очищаем список
    for i in CalendarID_group:
        try:
            now = datetime.utcnow().isoformat() + 'Z'
            events_result = service.events().list(calendarId=CalendarID_group[i], timeMin=now, maxResults=200, singleEvents=True, orderBy='startTime').execute()
            events = events_result.get('items', [])  # Ищем ивенты
            for event in events:  # Проходимся по найденным ивентам
                start = event['start'].get('dateTime')  # Записываем в переменную время ивента
                start = start.split('T')  # Разделяем на дату и время
                strs = event.get('id')  # Получаем id ивента
                if start[0] == time:  # если дата в ивенте == дате в планшетке
                    id_events.append(i)  # Записываем в список у какой группы мы парсили календарь
                    id_events.append(strs)  # добавляем в список id ивентов которые надо удалить
        except Exception as e:
            print(e)

    Name_list.clear()
    Delete_Events()

def planchet_id():
    """Ишем id планшеток по нынешной дате и запускае функцию ее скачивания"""
    try:
        response = requests.get(url='https://drive.google.com/drive/folders/1UH5pcJc0pxkYFWBMI_bNCxrdlApqv3nX')
        soup = BeautifulSoup(response.text, 'lxml')
        soup = soup.find_all('div', class_='Q5txwe')  # Достаем название планшетки

        for number, i in enumerate(soup, start=0):  # Проходимся по всем найденным планшеткам
            dates = str(datetime.now()).split()[0]
            dts = dates.split('-')
            dates = f'{dts[-1]}.{dts[1]}.{dts[0]}'  # Преобразуем текущую дату в нужный вид
            plans = i.text[0:-5]  # Убираем с названия планшетки '.xlsx'

            if dates == plans:  # Если текущая дата совпадает с датой планшетки
                response = requests.get(url='https://drive.google.com/drive/folders/1UH5pcJc0pxkYFWBMI_bNCxrdlApqv3nX')
                soup = BeautifulSoup(response.text, 'lxml')
                id_list = [i.replace('data-id="', '') for i in re.findall(r'data-id="[\w\-\+/#\(\)\*&\^:<>\?\!%\$]+', str(soup))]
                data_exel(identifier=id_list[number])  # Запускаем скачивание планшетки по ее id
                break
    except Exception as e:
        start_message(TG_Bot_ID, f'Была вызвана ошибка в разделе поиска  планшеток \n'
                                 f'описание ошибки: {e} \n'
                                 f'!ИЗ-ЗА ОШИБКИ ДАЛЬНЕЙШАЯ РАБОТА ПРОГРАММЫ ОСТАНОВЛЕНА ДО ПОВТОРНОГО ЕГО ВЫЗОВА!')

def data_exel(identifier: str):  # Функция скачивания планшетки
    """Функция скачивает найденную планшетку"""
    try:
        planchet = requests.get(url=f'https://drive.google.com/uc?export=download&id={identifier}')
        with open('planchette1.xlsx', 'wb') as xlsx_file:
            xlsx_file.write(planchet.content)
        planchette_check()  # Запускаем парсинг планшетки
    except Exception as e:
        start_message(TG_Bot_ID, f'''Была вызвана ошибка в разделе скачивания планшетки
                                 описание ошибки: {e}
                                 !ИЗ-ЗА ОШИБКИ ДАЛЬНЕЙШАЯ РАБОТА ПРОГРАММЫ ОСТАНОВЛЕНА ДО ПОВТОРНОГО ЕГО ВЫЗОВА!''')

def planchette_check():
    """Функция проверяет соответствует ли планшетка требованиям"""
    wb1 = openpyxl.load_workbook('planchette1.xlsx')
    worksheet = wb1['1 пара'].max_column
    if worksheet == 9:
        parsing()
    elif worksheet == 8:
        start_message(TG_Bot_ID, 'У планшетки отсувствует время пары, ошибка будет отработана')
        parsing()
    else:
        start_message(TG_Bot_ID, f'Планшетка не соответствует требованиям')

def parsing():
    wb1 = openpyxl.load_workbook('planchette1.xlsx')
    days = datetime.today().isoweekday()
    spisok = list()
    if days == 1:
        couple_list = ['1 пара', '2 пара', '3 пара', 'Классный час', '4 пара', '5 пара', '6 пара']
        par = ['Пара 1', 'Пара 2', 'Пара 3', 'Классный час', 'Пара 4', 'Пара 5', 'Пара 6']
    else:
        couple_list = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', '7 пара']
        par = ['Пара 1', 'Пара 2', 'Пара 3', 'Пара 4', 'Пара 5', 'Пара 6', 'Пара 7']


    try:
        for couple in couple_list:
            worksheet = wb1[couple]
            if worksheet.max_column == 9:
                dates = str(worksheet['B1'].value).split()
                if '-' in dates[0]:
                    datess = dates[0]
                else:
                    dates = dates[0].split('.')
                    datess = f'{dates[2]}-{dates[1]}-{dates[0]}'

                couple_time = worksheet['A1'].value
            elif worksheet.max_column == 8:
                dates = str(worksheet['A1'].value).split()
                if '-' in dates[0]:
                    datess = dates[0]
                else:
                    dates = dates[0].split('.')
                    datess = f'{dates[2]}-{dates[1]}-{dates[0]}'

                time1 = {'1 пара': '08:00-09:30', '2 пара': '09:40-11:10', '3 пара': '11:30-13:00',
                         '4 пара': '13:40-15:10', '5 пара': '15:30-17:00', '6 пара': '17:10-18:40',
                         'Классный час': '13:05-13:35'}
                time = {'1 пара': '08:00-09:30', '2 пара': '09:40-11:10', '3 пара': '11:30-13:00',
                        '4 пара': '13:10-14:40', '5 пара': '15:00-16:30', '6 пара': '16:40-18:10',
                        '7 пара': '18:20-19:50', 'Классный час': '13:05-13:35'}

                try:
                    if days == 1:
                        couple_time = time1[couple]
                    else:
                        couple_time = time[couple]
                except:
                    pass

            for j in range(0, worksheet.max_row):
                my_list_1 = []
                my_list_2 = []
                if worksheet.max_column == 9:
                    for col in worksheet.iter_cols(2, 5):
                        a = col[j].value
                        if a != None:
                            my_list_1.append(a)

                elif worksheet.max_column == 8:
                    for col in worksheet.iter_cols(1, 4):
                        a = col[j].value
                        if a != None:
                            my_list_1.append(a)

                if len(my_list_1) == 4:
                    if my_list_1[1] not in par:
                        if my_list_1[0] in CalendarID_group:
                            spisok.append([[my_list_1[0], my_list_1[1], my_list_1[2], my_list_1[-1]], [couple_time, datess]])
                        else:
                            my_list_1[0] = str(int(my_list_1[0]))
                            if my_list_1[0] in CalendarID_group:
                                spisok.append([[my_list_1[0], my_list_1[1], my_list_1[2], my_list_1[-1]], [couple_time, datess]])

                elif len(my_list_1) == 2:
                    if my_list_1[0] in CalendarID_group:
                        spisok.append([[my_list_1[0], 'Пусто', 'Пусто', my_list_1[-1]], [couple_time, datess]])
                    else:
                        my_list_1[0] = str(int(my_list_1[0]))
                        spisok.append([[my_list_1[0], 'Пусто', 'Пусто', my_list_1[-1]], [couple_time, datess]])

                if worksheet.max_column == 9:
                    for col in worksheet.iter_cols(6, 9):
                        a = col[j].value
                        if a != None:
                            my_list_2.append(a)

                elif worksheet.max_column == 8:
                    for col in worksheet.iter_cols(5, 8):
                        a = col[j].value
                        if a != None:
                            my_list_2.append(a)

                if len(my_list_2) == 4:
                    if my_list_2[1] not in par:
                        if my_list_2[0] in CalendarID_group:
                            spisok.append([[my_list_2[0], my_list_2[1], my_list_2[2], my_list_2[-1]], [couple_time, datess]])
                        else:
                            my_list_2[0] = str(int(my_list_2[0]))
                            if my_list_2[0] in CalendarID_group:
                                spisok.append([[my_list_2[0], my_list_2[1], my_list_2[2], my_list_2[-1]], [couple_time, datess]])
                elif len(my_list_2) == 2:
                    if my_list_2[0] in CalendarID_group:
                        spisok.append([[my_list_2[0], 'Пусто', 'Пусто', my_list_2[-1]], [couple_time, datess]])
                    else:
                        my_list_2[0] = str(int(my_list_2[0]))
                        spisok.append([[my_list_2[0], 'Пусто', 'Пусто', my_list_2[-1]], [couple_time, datess]])

        date = datetime.now()
        time = str(date.time()).split(':')
        Nowtime = f'{time[0]}:{time[1]}'
        date = date.date()
        get_events_list(str(date))

        for i in spisok:
            times = i[1][0].split('-')
            if Nowtime <= times[1]:
                New_Events(i[0][0], i[0][1], i[0][2], i[1][0], i[1][1], i[0][-1])
    except Exception as e:
        start_message(TG_Bot_ID, f'''Была вызвана ошибка в разделе парсинга планшетки(2.1)
                                 описание ошибки: {e}''')

def Par_Techer():  # Парсинг расписания преподователей
    """Функция достает расписание преподователей с сайта, обрабатывает и
    вызывает функцию для загрузки их в календарь"""
    url = 'https://rksi.ru/schedule'
    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'lxml')
    lecturer = soup.find('select', id='teacher').find_all('option')
    clear_list = []
    preps = []
    for i in lecturer:
        clear_list.append(i.text)  # Засовываем все группы в список

    for teacher in name_teacher:
        teac = f'{teacher[:-5]} {teacher[-5:]}'
        if teacher in clear_list or teac in clear_list:
            trash_prep_soup = ''
            if teacher in clear_list:
                trash_response_text = requests.post(url, {'teacher': f'{teacher}'.encode('cp1251'), 'stp': 'Показать!'.encode('cp1251')}).text
                trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')
            elif teac in clear_list:
                trash_response_text = requests.post(url, {'teacher': f'{teac}'.encode('cp1251'), 'stp': 'Показать!'.encode('cp1251')}).text
                trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')

            teachers_list = []
            for i in trash_prep_soup.find_all(['p', 'b']):
                teachers_list.append(
                    str(i).replace('<br/><b>', '||').replace('</b><br/>', '||').replace('<p>', '').replace('<b>', '').replace('</p>', '').replace('</b>', ''))

            teacher = teacher.replace('  ', ' ')
            preps.append(f'&{teacher}')
            for i in teachers_list:
                if '</' not in i:
                    preps.append(i.split('||'))
        break

    today = str(date.today())  # Текущая дата
    days = list()  # Засовываем дату (год-месяц-день) текущей пары
    groups = str()  # Для хранения группы
    for name in preps:  # Обрабатываем данные
        if name[0] == "&":
            groups = f"{name[1:]}"
        if 15 <= len(name[0]) <= 23 and name[0][0:1].isdigit():
            nd = name[0].split(" ")  # Разделяем дату на 'день', 'месяц', 'день недели'
            nd[1] = nd[1].replace(",", "")  # Избавляемся от знака , в месяце и засовываем в список
            month = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06',
                     'июля': '07', 'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12'}
            if len(nd[0]) == 1 and 1 <= int(nd[0]) < 10:
                days.append(f"{today[0:4]}-{month[nd[1]]}-0{nd[0]}")
            elif len(nd[0]) == 2 and int(nd[0]) >= 10:
                days.append(f"{today[0:4]}-{month[nd[1]]}-{nd[0]}")

        if len(name) == 3:  # Обработанные данные загружаем в календарь
            kab = name[2].split('ауд. ')[1].split('-1')[0]
            group = name[2].split(',')[0]
            New_Events(kab, group, groups, name[0], days[-1], name[1])

def Monday():
    for time in ['04:30', '07:15', '08:10', '09:00', '10:30', '12:20', '13:10', '15:00', '16:20', '17:45']:
        schedule.every().day.at(time).do(planchet_id)

def Other_days():
    for time in ['04:30', '07:15', '08:10', '09:00', '10:30', '12:20', '14:00', '15:50', '17:20', '18:34']:
        schedule.every().day.at(time).do(planchet_id)

def main():
    """Функция запускает другие функции в определенное время и день недели"""
    schedule.every().monday.at('04:25').do(Monday)
    schedule.every().tuesday.at('04:25').do(Other_days)
    schedule.every().wednesday.at('04:25').do(Other_days)
    schedule.every().thursday.at('04:25').do(Other_days)
    schedule.every().friday.at('04:25').do(Other_days)
    schedule.every().saturday.at('04:25').do(Other_days)
    schedule.every().sunday.at('05:00').do(clear_calendar)
    schedule.every().sunday.at('07:30').do(Par_Techer)
    schedule.every().sunday.at('09:30').do(Par_Techer)
    while True:
        schedule.run_pending()

if __name__ == '__main__':
    main()
    bot.polling()